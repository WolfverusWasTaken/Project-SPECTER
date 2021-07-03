# ==========================================================================================================
#                                            IMPORTED LIBRARIES                                            
# ==========================================================================================================
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import os

def data_addon(path, data_table_size, start_image, length_data, width_data, path_data, data_count):
    wb = load_workbook(path)
    sheet = wb.active

    data_table_size = 4
    image_table_size = 5

    iter = 31
    while (iter - 2) % (data_table_size + image_table_size) != 0:
        iter = iter + 1
    iter += 1

    total_data = (iter - 3) / (data_table_size + image_table_size)

    xmax = data_count/10
    if xmax.is_integer() == False:
        xmax = int(xmax) + 1

    rowpos = 1

    '''for x in range(2, iter, data_table_size + image_table_size):
        for y in range(2, 7, 1):
            sheet.cell(row=x + 5, column=y).value =  length_data[y-2]
            sheet.cell(row=x + 6, column=y).value =  width_data[y-2]'''

    for y in range(2, data_count + 1, 1):

        if (((y - 2) / 10).is_integer()):
            if rowpos < 2:
                rowpos += 1
            else:
                rowpos += 9

            minlist = y - 2
            maxlist = y - 2 + 9
            # print("Row :", rowpos, ":", y-2, "to", y-2+9)

            for colpos in range(minlist, maxlist + 1, 1):
                if colpos < data_count - 1:

                    colpos = colpos % 10

                    char = number2text(colpos + 2)
                    #print("Row :", rowpos, ":", colpos, ":", char)
                    #print(colpos,":" ,char)
                    cellval = char + str(rowpos)

                    print(cellval + " : " + path_data[colpos])
                    sheet.cell(row=rowpos + 5, column=colpos + 2).value = length_data[colpos]
                    sheet.cell(row=rowpos + 6, column=colpos + 2).value = width_data[colpos]
                    img = openpyxl.drawing.image.Image(path_data[colpos])

                    img.height = 96
                    img.width = 64
                    sheet.add_image(img, char + str(rowpos))

    print("done")

    wb.save(path)

def addon_template_image(filepath,imgpath, data_table_size, image_table_size, image_start_col, start_image):
    wb = openpyxl.Workbook()
    sheet = wb.active

    data_table_size = 4
    image_table_size = 5
    image_start_col = 2
    start_image = 10.0

    iter = 31

    img = openpyxl.drawing.image.Image('Capture2.PNG')

    while (iter - 2) % (data_table_size + image_table_size) != 0:
        iter = iter + 1

    iter += 1

    total_data = (iter - 3) / (data_table_size + image_table_size)

    for i in range(2, iter, data_table_size + image_table_size):

        for file in os.listdir(imgpath):
            if file.endswith(".PNG"):
                img = openpyxl.drawing.image.Image(os.path.join(imgpath, file))
            continue

        '''sheet.merge_cells(start_row=i,
                          start_column=image_start_col,
                          end_row=(i + image_table_size) - 1,
                          end_column=image_start_col + image_table_size - 1)

        sheet.merge_cells(start_row=i,
                          start_column=image_start_col + image_table_size,
                          end_row=(i + image_table_size) - 1,
                          end_column=image_start_col + image_table_size + image_table_size - 1)'''

        for x in range(2, 12, 1):
            sheet.merge_cells(start_row=i,
                              start_column=x,
                              end_row=i + image_table_size - 1,
                              end_column=x)


        sheet.cell(row=i - 1, column=image_start_col - 1).value = "Data: "
        sheet.cell(row=i, column=image_start_col - 1).value = "Img: "
        sheet.cell(row=i + 5, column=image_start_col - 1).value = "Length: "
        sheet.cell(row=i + 6, column=image_start_col - 1).value = "Width: "

        for x in range(1, 11):
            sheet.cell(row=i - 1, column=x + 1).value = float(start_image + (x - 1) / 10)


        start_image = start_image + 1

    wb.save(filepath)

def number2text(number):
    if number == 1:
        return "A"
    if number == 2:
        return "B"
    if number == 3:
        return "C"
    if number == 4:
        return "D"
    if number == 5:
        return "E"
    if number == 6:
        return "F"
    if number == 7:
        return "G"
    if number == 8:
        return "H"
    if number == 9:
        return "I"
    if number == 10:
        return "J"
    if number == 11:
        return "K"